# /// script
# requires-python = ">=3.12"
# dependencies = [
#    "rich",
#    "openpyxl",
#    "pandas",
# ]
# ///

import sys
import re
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
from rich import print

def main():
    if len(sys.argv) != 2:
        print("[red]Usage: python script.py <excel_file>[/red]")
        sys.exit(1)

    input_file = Path(sys.argv[1])
    if not input_file.exists():
        print(f"[red]Error: File {input_file} does not exist[/red]")
        sys.exit(1)

    output_file = input_file.stem + "_output.csv"
    data = []

    workbook = load_workbook(input_file, data_only=True)
    for year in range(1999, 2024):
        sheet_name = str(year)
        if sheet_name not in workbook.sheetnames:
            print(f"[yellow]Warning: Sheet for {year} not found, skipping...[/yellow]")
            continue

        sheet = workbook[sheet_name]
        labels = [sheet[f"A{row}"].value for row in range(12, 70)]
        switzerland_values = [sheet[f"C{row}"].value for row in range(12, 70)]
        netherlands_values = [sheet[f"W{row}"].value for row in range(12, 70)]

        # Use regex to remove numbers and parentheses in labels
        labels = [re.sub(r"\s*\d+\)$", "", label) if label else label for label in labels]

        for label, ch_value, nl_value in zip(labels, switzerland_values, netherlands_values):
            if ch_value is not None and nl_value is not None:
                data.append({"Year": year, "Label": label, "Switzerland": ch_value, "Netherlands": nl_value})

    df = pd.DataFrame(data)
    df.to_csv(output_file, sep=";", index=False)
    print(f"[green]Data successfully saved to {output_file}[/green]")

if __name__ == "__main__":
    main()
